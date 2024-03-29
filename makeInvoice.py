import openpyxl
from openpyxl import load_workbook
from openpyxl import drawing
from openpyxl.styles import PatternFill

def makeFile(templatePath, outputPath):
    """
    makeFile makes an Excel book which copies the template excel file

    :param templatePath: string - absolute path of the template excel file
    :param outputPath: string - new Excel book
    :return: void
    """
    workbook = openpyxl.load_workbook(templatePath)
    workbook.save(outputPath)

def pasteLogo(sheet):
    """
    pasteLogo is a void function which pastes the logo at the top of the invoice

    :param sheet: Excel Sheet Object
    :return: void
    """
    # Make image object
    logo = drawing.image.Image("/Users/claudio/Desktop/CS_Projects/Automated_Invoice/lorhLogo.png")
    logo.width = 850
    logo.height = 135
    sheet.add_image(logo, "A1") # Insert image object (size already adjusted)


def fillBaseInfo(sheet, baseInfoList):
    """
    fillBaseInfo simply fills data in upper side of the invoice

    :param sheet: Excel sheet object
    :param baseInfoList: list - containing client data (name, address and invoice matter)
    :return: void
    """
    # Matter is stored in the last and fill in cell C8
    matterStr = baseInfoList.pop()
    sheet.cell(row=8, column=3, value=matterStr)

    for rowNum, value in enumerate(baseInfoList, start=7):
        sheet.cell(row=rowNum, column=5, value=value)

def findEndRow(sheet, rowNum):
    """
    findEndRow finds a row number which is not empty

    :param sheet: Sheet Object
    :param rowNum: int - from which row to start (this is the end of the entry of the work activities)
    :return: int - row number
    """
    columnB = sheet.iter_cols(min_col=2, max_col=2, # Only qualify column B
                              min_row=rowNum, values_only=True)
    
    for column in columnB:
        for value in column:
            if value is not None:
                return rowNum
            rowNum += 1

    return None

def fillFormula(sheet, rowNum):
    """
    fillFormula is a void function which fills some Excel formulas in cells

    :param sheet: Sheet Object to be filled formula
    :param rowNum: int - from which row the formula should be filled
    :return: void
    """
    # Managing Attorney HOURS and AMOUNT
    sheet[f"D{rowNum}"] = f"=SUMIF(E15:E{rowNum-1}, \">=485\", D15:D{rowNum-1})"
    sheet[f"F{rowNum}"] = f"=SUMIF(E15:E{rowNum-1}, \">=485\", F15:F{rowNum-1})"

    # Associate Attorney HOURS and AMOUNT
    sheet[f"D{rowNum+1}"] = f"=SUMIFS(D15:D{rowNum-1}, E15:E{rowNum-1}, \">=250\", E15:E{rowNum-1}, \"<485\")"
    sheet[f"F{rowNum+1}"] = f"=SUMIFS(F15:F{rowNum-1}, E15:E{rowNum-1}, \">=250\", E15:E{rowNum-1}, \"<485\")"

    # Law Cleak/Paralegal HOURS and AMOUNT
    sheet[f"D{rowNum+2}"] = f"=SUMIFS(D15:D{rowNum-1}, E15:E{rowNum-1}, \">=200\", E15:E{rowNum-1}, \"<250\")"
    sheet[f"F{rowNum+2}"] = f"=SUMIFS(F15:F{rowNum-1}, E15:E{rowNum-1}, \">=200\", E15:E{rowNum-1}, \"<250\")"

    # Legal Assistant HOURS and AMOUNT
    sheet[f"D{rowNum+3}"] = f"=SUMIF(E15:E{rowNum-1}, \"<200\", D15:D{rowNum-1})"
    sheet[f"F{rowNum+3}"] = f"=SUMIF(E15:E{rowNum-1}, \"<200\", F15:F{rowNum-1})"

    # Total Attorneys' Fees:
    sheet[f"F{rowNum+4}"] = f"=SUM(F{rowNum}:F{rowNum+3})"

    # TOTAL AMOUNT CURRENTLY DUE:
    sheet[f"F{rowNum+16}"] = f"=SUM(F{rowNum+4}, F{rowNum+7}:F{rowNum+8}, F{rowNum+11}:F{rowNum+12})"

def makeInvoice(outputPath, baseInfoDict, dataDict, invDate):
    """
    makeInvoice makes invoice sheets as an Excel file.

    :param outputPath: string - absolute path for the Excel file
    :param baseInfoDict: dictionary - containing base information of each client
    :param dataDict: dictionary - containing invoicing data (date, work, rate and time)
    :param invDate: string - invoice issuing date
    :return: void
    """
    workbook = load_workbook(outputPath)
    templateSheet = workbook.active

    for client, entryList in dataDict.items():
        newSheet = workbook.copy_worksheet(templateSheet)

        pasteLogo(newSheet)

        # Name sheet with client name as appears in ACR
        newSheet.title = client

        # Fill invoice issue date
        newSheet.cell(row=7, column=3, value=invDate)

        # baseInfo is Bill To and Matter section. Matter is the last in the list.
        baseInfoList = baseInfoDict[client]
        fillBaseInfo(newSheet, baseInfoList)

        lightGrey = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

        rowNum = 15 # Start row. Need to specify!
        for rowData in entryList:
            for colNum, value in enumerate(rowData, 2): # Start column. Need to specify!
                cell = newSheet.cell(row=rowNum, column=colNum, value=value)
                if rowNum%2 == 0: # If row # is even, fill lightgrey in cells
                    cell.fill = lightGrey
            
            timeValue = newSheet[f"D{rowNum}"].value
            if timeValue != "Flat Fee" and timeValue != "Not Billed":
                cell = newSheet.cell(row=rowNum, column=6, value=f"=D{rowNum}*E{rowNum}")
            else:
                cell = newSheet.cell(row=rowNum, column=6, value=0)
            if rowNum%2 == 0:
                cell.fill = lightGrey

            rowNum += 1
        
        endRow = findEndRow(newSheet, rowNum) # Want to find the first row which is not empty
        numToDelete = endRow - rowNum + 1

        for _ in range(numToDelete):
             newSheet.delete_rows(rowNum)

        fillFormula(newSheet, rowNum)

    workbook.save(outputPath)
